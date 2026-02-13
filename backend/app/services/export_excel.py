"""Excel export functions for financial reports using openpyxl."""
from __future__ import annotations

import io
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from backend.app.services.export_i18n import t

# ── Shared styling constants ────────────────────────────────────────────────

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_SECTION_FONT = Font(name="Calibri", bold=True, size=11)
_SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_TOTAL_FONT = Font(name="Calibri", bold=True, size=11)
_TOTAL_BORDER = Border(
    top=Side(style="thin"),
    bottom=Side(style="double"),
)
_CURRENCY_FMT = '#,##0.0000'
_RIGHT = Alignment(horizontal="right")
_LEFT = Alignment(horizontal="left")


def _auto_width(ws: Any) -> None:
    """Auto-fit column widths based on content."""
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, values_only=False):
            cell = row[0]
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)


def _write_header_row(ws: Any, row: int, values: list[str]) -> None:
    """Write a styled header row."""
    for col, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _RIGHT if col > 1 else _LEFT


def _write_title(ws: Any, title: str, subtitle: str) -> int:
    """Write report title and subtitle, return next available row."""
    ws.cell(row=1, column=1, value=title).font = Font(name="Calibri", bold=True, size=14)
    ws.cell(row=2, column=1, value=subtitle).font = Font(name="Calibri", size=10, italic=True)
    return 4


def _to_workbook(ws: Any, wb: Workbook) -> io.BytesIO:
    """Finalize workbook and return as BytesIO."""
    _auto_width(ws)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── 1. Income Statement ────────────────────────────────────────────────────


def export_income_statement_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "income_statement")

    row = _write_title(ws, t(lang, "income_statement"), f"{t(lang, 'period')}: {data['from_date']} to {data['to_date']}")

    # Revenue section
    ws.cell(row=row, column=1, value=t(lang, "revenue")).font = _SECTION_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    row += 1
    for item in data.get("revenue_detail", []):
        ws.cell(row=row, column=1, value=f"  {item['name']}")
        c = ws.cell(row=row, column=2, value=float(item["amount"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_revenue")).font = _TOTAL_FONT
    c = ws.cell(row=row, column=2, value=float(data["revenue"]))
    c.number_format = _CURRENCY_FMT
    c.font = _TOTAL_FONT
    c.border = _TOTAL_BORDER
    c.alignment = _RIGHT
    row += 2

    # COGS section
    ws.cell(row=row, column=1, value=t(lang, "cogs")).font = _SECTION_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    row += 1
    for item in data.get("expense_detail", []):
        if item["code"] == "5000":
            ws.cell(row=row, column=1, value=f"  {item['name']}")
            c = ws.cell(row=row, column=2, value=float(item["amount"]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
            row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_cogs")).font = _TOTAL_FONT
    c = ws.cell(row=row, column=2, value=float(data["cogs"]))
    c.number_format = _CURRENCY_FMT
    c.font = _TOTAL_FONT
    c.border = _TOTAL_BORDER
    c.alignment = _RIGHT
    row += 2

    # Gross Profit
    ws.cell(row=row, column=1, value=t(lang, "gross_profit")).font = Font(name="Calibri", bold=True, size=12)
    c = ws.cell(row=row, column=2, value=float(data["gross_profit"]))
    c.number_format = _CURRENCY_FMT
    c.font = Font(name="Calibri", bold=True, size=12)
    c.alignment = _RIGHT
    row += 2

    # Operating Expenses
    ws.cell(row=row, column=1, value=t(lang, "operating_expenses")).font = _SECTION_FONT
    ws.cell(row=row, column=1).fill = _SECTION_FILL
    ws.cell(row=row, column=2).fill = _SECTION_FILL
    row += 1
    for item in data.get("expense_detail", []):
        if item["code"] != "5000":
            ws.cell(row=row, column=1, value=f"  {item['name']}")
            c = ws.cell(row=row, column=2, value=float(item["amount"]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
            row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_opex")).font = _TOTAL_FONT
    c = ws.cell(row=row, column=2, value=float(data["operating_expenses"]))
    c.number_format = _CURRENCY_FMT
    c.font = _TOTAL_FONT
    c.border = _TOTAL_BORDER
    c.alignment = _RIGHT
    row += 2

    # Net Income
    ws.cell(row=row, column=1, value=t(lang, "net_income")).font = Font(name="Calibri", bold=True, size=13)
    c = ws.cell(row=row, column=2, value=float(data["net_income"]))
    c.number_format = _CURRENCY_FMT
    c.font = Font(name="Calibri", bold=True, size=13)
    c.border = Border(top=Side(style="double"), bottom=Side(style="double"))
    c.alignment = _RIGHT

    return _to_workbook(ws, wb)


# ── 2. Trial Balance ──────────────────────────────────────────────────────


def export_trial_balance_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "trial_balance")

    row = _write_title(ws, t(lang, "trial_balance"), f"{t(lang, 'period')}: {data['from_date']} to {data['to_date']}")
    _write_header_row(ws, row, [t(lang, "code"), t(lang, "account"), t(lang, "type"), t(lang, "debit"), t(lang, "credit")])
    row += 1

    for acct in data.get("accounts", []):
        ws.cell(row=row, column=1, value=acct["account_code"])
        ws.cell(row=row, column=2, value=acct["account_name"])
        ws.cell(row=row, column=3, value=acct["account_type"])
        c = ws.cell(row=row, column=4, value=float(acct["debit"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=row, column=5, value=float(acct["credit"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        row += 1

    # Totals
    ws.cell(row=row, column=1, value=t(lang, "totals")).font = _TOTAL_FONT
    for col, key in [(4, "total_debit"), (5, "total_credit")]:
        c = ws.cell(row=row, column=col, value=float(data[key]))
        c.number_format = _CURRENCY_FMT
        c.font = _TOTAL_FONT
        c.border = _TOTAL_BORDER
        c.alignment = _RIGHT

    row += 1
    balanced = t(lang, "balanced") if data.get("is_balanced") else t(lang, "out_of_balance")
    ws.cell(row=row, column=1, value=balanced).font = Font(
        name="Calibri", bold=True, color="008000" if data.get("is_balanced") else "FF0000",
    )

    return _to_workbook(ws, wb)


# ── 3. Balance Sheet ──────────────────────────────────────────────────────


def export_balance_sheet_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "balance_sheet")

    row = _write_title(ws, t(lang, "balance_sheet"), f"{t(lang, 'as_of')} {data['as_of_date']}")

    section_map = [
        ("assets", "assets", "total_assets"),
        ("liabilities", "liabilities", "total_liabilities"),
        ("equity", "equity", "total_equity"),
    ]
    total_label_map = {
        "assets": "total_assets",
        "liabilities": "total_liabilities",
        "equity": "total_equity",
    }

    for section_key, items_key, total_key in section_map:
        ws.cell(row=row, column=1, value=t(lang, section_key)).font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _SECTION_FILL
        ws.cell(row=row, column=2).fill = _SECTION_FILL
        ws.cell(row=row, column=3).fill = _SECTION_FILL
        row += 1
        _write_header_row(ws, row, [t(lang, "code"), t(lang, "account"), t(lang, "balance")])
        row += 1
        for item in data.get(items_key, []):
            ws.cell(row=row, column=1, value=item["code"])
            ws.cell(row=row, column=2, value=item["name"])
            c = ws.cell(row=row, column=3, value=float(item["balance"]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
            row += 1

        if section_key == "equity":
            ws.cell(row=row, column=2, value=t(lang, "retained_earnings"))
            c = ws.cell(row=row, column=3, value=float(data["retained_earnings"]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
            row += 1

        ws.cell(row=row, column=1, value=t(lang, total_label_map[section_key])).font = _TOTAL_FONT
        c = ws.cell(row=row, column=3, value=float(data[total_key]))
        c.number_format = _CURRENCY_FMT
        c.font = _TOTAL_FONT
        c.border = _TOTAL_BORDER
        c.alignment = _RIGHT
        row += 2

    # Total L&E
    ws.cell(row=row, column=1, value=t(lang, "total_le")).font = Font(name="Calibri", bold=True, size=12)
    c = ws.cell(row=row, column=3, value=float(data["total_liabilities_and_equity"]))
    c.number_format = _CURRENCY_FMT
    c.font = Font(name="Calibri", bold=True, size=12)
    c.border = Border(top=Side(style="double"), bottom=Side(style="double"))
    c.alignment = _RIGHT

    row += 1
    balanced = t(lang, "balanced") if data.get("is_balanced") else t(lang, "out_of_balance")
    ws.cell(row=row, column=1, value=balanced).font = Font(
        name="Calibri", bold=True, color="008000" if data.get("is_balanced") else "FF0000",
    )

    return _to_workbook(ws, wb)


# ── 4. General Ledger ────────────────────────────────────────────────────


def export_general_ledger_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "general_ledger")

    row = _write_title(
        ws,
        t(lang, "general_ledger"),
        f"{t(lang, 'account')}: {data['account_code']} \u2014 {data['account_name']}  |  {data['from_date']} to {data['to_date']}",
    )

    _write_header_row(ws, row, [t(lang, "date"), t(lang, "reference"), t(lang, "description"), t(lang, "debit"), t(lang, "credit"), t(lang, "balance")])
    row += 1

    # Opening balance
    ws.cell(row=row, column=1, value=t(lang, "opening_balance")).font = _SECTION_FONT
    c = ws.cell(row=row, column=6, value=float(data["opening_balance"]))
    c.number_format = _CURRENCY_FMT
    c.font = _SECTION_FONT
    c.alignment = _RIGHT
    row += 1

    for entry in data.get("entries", []):
        ws.cell(row=row, column=1, value=entry["date"])
        ws.cell(row=row, column=2, value=entry.get("reference") or "")
        ws.cell(row=row, column=3, value=entry["description"])
        c = ws.cell(row=row, column=4, value=float(entry["debit"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=row, column=5, value=float(entry["credit"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=row, column=6, value=float(entry["running_balance"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        row += 1

    # Closing balance
    ws.cell(row=row, column=1, value=t(lang, "closing_balance")).font = _TOTAL_FONT
    c = ws.cell(row=row, column=6, value=float(data["closing_balance"]))
    c.number_format = _CURRENCY_FMT
    c.font = _TOTAL_FONT
    c.border = _TOTAL_BORDER
    c.alignment = _RIGHT

    return _to_workbook(ws, wb)


# ── 5. VAT Report ────────────────────────────────────────────────────────


def export_vat_report_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "vat_report")

    row = _write_title(ws, t(lang, "vat_report"), f"{t(lang, 'period')}: {data['from_date']} to {data['to_date']}")

    # Summary
    ws.cell(row=row, column=1, value=t(lang, "total_vat_collected")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(data["total_vat_collected"]))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_sales_ex_vat")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(data["total_sales_ex_vat"]))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 1
    ws.cell(row=row, column=1, value=t(lang, "effective_vat_rate")).font = _SECTION_FONT
    ws.cell(row=row, column=2, value=f"{data['effective_vat_rate']}%").alignment = _RIGHT
    row += 2

    # Monthly breakdown
    _write_header_row(ws, row, [t(lang, "month"), t(lang, "vat_collected"), t(lang, "sales_ex_vat"), t(lang, "transactions")])
    row += 1
    for m in data.get("monthly_breakdown", []):
        ws.cell(row=row, column=1, value=m["month"])
        c = ws.cell(row=row, column=2, value=float(m["vat_collected"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=row, column=3, value=float(m["sales_ex_vat"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        ws.cell(row=row, column=4, value=m["transaction_count"]).alignment = _RIGHT
        row += 1

    return _to_workbook(ws, wb)


# ── 6. Cash Flow ─────────────────────────────────────────────────────────


def export_cash_flow_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "cash_flow")

    row = _write_title(ws, t(lang, "cash_flow"), f"{t(lang, 'period')}: {data['from_date']} to {data['to_date']}")

    # Opening balance
    ws.cell(row=row, column=1, value=t(lang, "opening_cash")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(data["opening_cash_balance"]))
    c.number_format = _CURRENCY_FMT
    c.font = _SECTION_FONT
    c.alignment = _RIGHT
    row += 2

    for section_label_key, section_key in [
        ("operating", "operating"),
        ("investing", "investing"),
        ("financing", "financing"),
    ]:
        section = data.get(section_key, {})
        ws.cell(row=row, column=1, value=t(lang, section_label_key)).font = _SECTION_FONT
        ws.cell(row=row, column=1).fill = _SECTION_FILL
        ws.cell(row=row, column=2).fill = _SECTION_FILL
        row += 1
        for item in section.get("items", []):
            ws.cell(row=row, column=1, value=f"  {item['description']}")
            c = ws.cell(row=row, column=2, value=float(item["amount"]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
            row += 1
        ws.cell(row=row, column=1, value=f"{t(lang, 'total')} {t(lang, section_label_key)}").font = _TOTAL_FONT
        c = ws.cell(row=row, column=2, value=float(section.get("total", "0")))
        c.number_format = _CURRENCY_FMT
        c.font = _TOTAL_FONT
        c.border = _TOTAL_BORDER
        c.alignment = _RIGHT
        row += 2

    # Net change
    ws.cell(row=row, column=1, value=t(lang, "net_change")).font = Font(name="Calibri", bold=True, size=12)
    c = ws.cell(row=row, column=2, value=float(data["net_change"]))
    c.number_format = _CURRENCY_FMT
    c.font = Font(name="Calibri", bold=True, size=12)
    c.alignment = _RIGHT
    row += 2

    # Closing balance
    ws.cell(row=row, column=1, value=t(lang, "closing_cash")).font = Font(name="Calibri", bold=True, size=13)
    c = ws.cell(row=row, column=2, value=float(data["closing_cash_balance"]))
    c.number_format = _CURRENCY_FMT
    c.font = Font(name="Calibri", bold=True, size=13)
    c.border = Border(top=Side(style="double"), bottom=Side(style="double"))
    c.alignment = _RIGHT

    return _to_workbook(ws, wb)


# ── 7. AR Aging ──────────────────────────────────────────────────────────


def export_ar_aging_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "ar_aging")

    row = _write_title(ws, t(lang, "ar_aging"), f"{t(lang, 'as_of')} {data['as_of_date']}")

    # KPI
    kpi = data.get("kpi", {})
    ws.cell(row=row, column=1, value=t(lang, "total_receivable")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(kpi.get("total_receivable", "0")))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_overdue")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(kpi.get("total_overdue", "0")))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 1
    ws.cell(row=row, column=1, value=t(lang, "dso")).font = _SECTION_FONT
    ws.cell(row=row, column=2, value=f"{kpi.get('dso', '0')} days").alignment = _RIGHT
    row += 2

    # Table
    _write_header_row(ws, row, [t(lang, "customer"), t(lang, "current_0_30"), t(lang, "days_31_60"), t(lang, "days_61_90"), t(lang, "over_90"), t(lang, "total")])
    row += 1
    for cust in data.get("customers", []):
        ws.cell(row=row, column=1, value=cust["name"])
        for col, key in [(2, "current"), (3, "days_31_60"), (4, "days_61_90"), (5, "over_90"), (6, "total")]:
            c = ws.cell(row=row, column=col, value=float(cust[key]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
        row += 1

    # Totals
    totals = data.get("totals", {})
    ws.cell(row=row, column=1, value=t(lang, "total")).font = _TOTAL_FONT
    for col, key in [(2, "current"), (3, "days_31_60"), (4, "days_61_90"), (5, "over_90"), (6, "total")]:
        c = ws.cell(row=row, column=col, value=float(totals.get(key, "0")))
        c.number_format = _CURRENCY_FMT
        c.font = _TOTAL_FONT
        c.border = _TOTAL_BORDER
        c.alignment = _RIGHT

    return _to_workbook(ws, wb)


# ── 8. AP Aging ──────────────────────────────────────────────────────────


# ── 9. Inventory Valuation ──────────────────────────────────────────────


def export_inventory_valuation_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "inventory_valuation")

    subtitle = f"{t(lang, 'as_of')} {data['as_of_date']}"
    if data.get("warehouse_filter"):
        subtitle += f"  |  {t(lang, 'warehouse')}: {data['warehouse_filter']}"
    if data.get("category_filter"):
        subtitle += f"  |  {t(lang, 'category')}: {data['category_filter']}"
    row = _write_title(ws, t(lang, "inventory_valuation"), subtitle)

    _write_header_row(ws, row, [t(lang, "sku"), t(lang, "product"), t(lang, "category"), t(lang, "quantity"), t(lang, "cost_price"), t(lang, "total_value")])
    row += 1

    for item in data.get("items", []):
        ws.cell(row=row, column=1, value=item["sku"])
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["category"])
        ws.cell(row=row, column=4, value=item["quantity"]).alignment = _RIGHT
        c = ws.cell(row=row, column=5, value=float(item["cost_price"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        c = ws.cell(row=row, column=6, value=float(item["total_value"]))
        c.number_format = _CURRENCY_FMT
        c.alignment = _RIGHT
        row += 1

    # Summary row
    ws.cell(row=row, column=1, value=t(lang, "totals")).font = _TOTAL_FONT
    ws.cell(row=row, column=3, value=f"{data['total_items']} {t(lang, 'items')}").font = _TOTAL_FONT
    c = ws.cell(row=row, column=4, value=data["total_quantity"])
    c.font = _TOTAL_FONT
    c.alignment = _RIGHT
    c = ws.cell(row=row, column=6, value=float(data["total_value"]))
    c.number_format = _CURRENCY_FMT
    c.font = _TOTAL_FONT
    c.border = _TOTAL_BORDER
    c.alignment = _RIGHT

    return _to_workbook(ws, wb)


def export_ap_aging_excel(data: dict[str, Any], lang: str = "en") -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = t(lang, "ap_aging")

    row = _write_title(ws, t(lang, "ap_aging"), f"{t(lang, 'as_of')} {data['as_of_date']}")

    # KPI
    kpi = data.get("kpi", {})
    ws.cell(row=row, column=1, value=t(lang, "total_payable")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(kpi.get("total_payable", "0")))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 1
    ws.cell(row=row, column=1, value=t(lang, "total_overdue")).font = _SECTION_FONT
    c = ws.cell(row=row, column=2, value=float(kpi.get("total_overdue", "0")))
    c.number_format = _CURRENCY_FMT
    c.alignment = _RIGHT
    row += 2

    # Table
    _write_header_row(ws, row, [t(lang, "supplier"), t(lang, "current_0_30"), t(lang, "days_31_60"), t(lang, "days_61_90"), t(lang, "over_90"), t(lang, "total")])
    row += 1
    for supp in data.get("suppliers", []):
        ws.cell(row=row, column=1, value=supp["name"])
        for col, key in [(2, "current"), (3, "days_31_60"), (4, "days_61_90"), (5, "over_90"), (6, "total")]:
            c = ws.cell(row=row, column=col, value=float(supp[key]))
            c.number_format = _CURRENCY_FMT
            c.alignment = _RIGHT
        row += 1

    # Totals
    totals = data.get("totals", {})
    ws.cell(row=row, column=1, value=t(lang, "total")).font = _TOTAL_FONT
    for col, key in [(2, "current"), (3, "days_31_60"), (4, "days_61_90"), (5, "over_90"), (6, "total")]:
        c = ws.cell(row=row, column=col, value=float(totals.get(key, "0")))
        c.number_format = _CURRENCY_FMT
        c.font = _TOTAL_FONT
        c.border = _TOTAL_BORDER
        c.alignment = _RIGHT

    return _to_workbook(ws, wb)
